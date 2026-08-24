"""Excel template + import User Stories (kèm AC) / Permission Matrix."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


def build_template_bytes() -> bytes:
    wb = Workbook()
    header_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="1A1A1A"),
        right=Side(style="thin", color="1A1A1A"),
        top=Side(style="thin", color="1A1A1A"),
        bottom=Side(style="thin", color="1A1A1A"),
    )
    body_align = Alignment(vertical="center", wrap_text=True)

    def style_header_row(ws, headers: list[str]):
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(1, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin
        ws.row_dimensions[1].height = 28

    # --- UserStories ---
    ws = wb.active
    ws.title = "UserStories"
    style_header_row(ws, ["content", "acceptance_criteria"])
    samples = [
        [
            "As a Sales Staff, I want to view assigned customers so that I can manage my accounts.",
            "System shows only assigned customers.;Staff must be logged in.",
        ],
        [
            "As an Admin, I want to delete user accounts so that I can remove inactive users.",
            "Only inactive accounts can be deleted.",
        ],
    ]
    for row in samples:
        ws.append(row)
        for col_idx in range(1, 3):
            c = ws.cell(ws.max_row, col_idx)
            c.alignment = body_align
            c.border = thin
    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 55

    # --- PermissionMatrix ---
    ws2 = wb.create_sheet("PermissionMatrix")
    headers_pm = ["role", "action", "resource", "effect", "scope", "condition"]
    style_header_row(ws2, headers_pm)
    pm_samples = [
        ["Sales Staff", "view", "Customer", "Allow", "assigned", ""],
        ["Admin", "delete", "User Account", "Allow", "all", ""],
    ]
    for row in pm_samples:
        ws2.append(row)
        for col_idx in range(1, 7):
            c = ws2.cell(ws2.max_row, col_idx)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin
    for col, w in zip("ABCDEF", [18, 12, 18, 10, 12, 28]):
        ws2.column_dimensions[col].width = w

    # --- README ---
    ws3 = wb.create_sheet("README")
    ws3.append(["Huong dan import ReqSentinel"])
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.append([])
    ws3.append(["Sheet UserStories"])
    ws3.append(["- content: bat buoc, 1 user story moi dong"])
    ws3.append(["- acceptance_criteria: tuy chon; nhieu AC cach nhau bang dau ;"])
    ws3.append([])
    ws3.append(["Sheet PermissionMatrix"])
    ws3.append(["- role, action, resource: bat buoc"])
    ws3.append(["- effect: Allow hoac Deny (mac dinh Allow)"])
    ws3.append(["- scope, condition: tuy chon"])
    ws3.column_dimensions["A"].width = 90

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_import_excel(content: bytes) -> dict[str, Any]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    errors: list[str] = []
    user_stories: list[dict] = []
    pm_rows: list[dict] = []

    if "UserStories" in wb.sheetnames:
        ws = wb["UserStories"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
            try:
                i_content = header.index("content")
            except ValueError:
                errors.append("UserStories thieu cot content")
                i_content = -1
            i_ac = header.index("acceptance_criteria") if "acceptance_criteria" in header else -1

            if i_content >= 0:
                for n, row in enumerate(rows[1:], start=2):
                    if not row or all(c is None or str(c).strip() == "" for c in row):
                        continue
                    content = row[i_content] if i_content < len(row) else None
                    if content is None or not str(content).strip():
                        errors.append(f"UserStories dong {n}: content trong — bo qua")
                        continue
                    acs: list[str] = []
                    if i_ac >= 0 and i_ac < len(row) and row[i_ac] is not None:
                        raw = str(row[i_ac]).strip()
                        if raw:
                            acs = [
                                p.strip()
                                for p in raw.replace("\n", ";").split(";")
                                if p.strip()
                            ]
                    user_stories.append(
                        {
                            "content": str(content).strip(),
                            "acceptance_criteria": acs,
                        }
                    )
        else:
            errors.append("Sheet UserStories trong")
    else:
        errors.append("Thieu sheet UserStories")

    if "PermissionMatrix" in wb.sheetnames:
        ws = wb["PermissionMatrix"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

            def col(name: str) -> int:
                return header.index(name) if name in header else -1

            i_role, i_action, i_res = col("role"), col("action"), col("resource")
            i_effect, i_scope, i_cond = col("effect"), col("scope"), col("condition")

            if min(i_role, i_action, i_res) < 0:
                errors.append("PermissionMatrix thieu role/action/resource")
            else:
                for n, row in enumerate(rows[1:], start=2):
                    if not row or all(c is None or str(c).strip() == "" for c in row):
                        continue

                    def cell(i: int) -> str | None:
                        if i < 0 or i >= len(row) or row[i] is None:
                            return None
                        s = str(row[i]).strip()
                        return s or None

                    role, action, resource = cell(i_role), cell(i_action), cell(i_res)
                    if not role or not action or not resource:
                        errors.append(
                            f"PermissionMatrix dong {n}: thieu role/action/resource — bo qua"
                        )
                        continue
                    effect = (cell(i_effect) or "Allow")
                    effect = effect[:1].upper() + effect[1:].lower() if effect else "Allow"
                    if effect not in ("Allow", "Deny"):
                        effect = "Allow"
                    pm_rows.append(
                        {
                            "role": role,
                            "action": action,
                            "resource": resource,
                            "effect": effect,
                            "scope": cell(i_scope),
                            "condition": cell(i_cond),
                        }
                    )
        else:
            errors.append("Sheet PermissionMatrix trong")
    else:
        errors.append("Thieu sheet PermissionMatrix")

    wb.close()
    return {
        "user_stories": user_stories,
        "permission_matrix": pm_rows,
        "errors": errors,
    }